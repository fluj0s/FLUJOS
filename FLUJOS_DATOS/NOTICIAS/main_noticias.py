import os
import re
import hashlib
import requests
import time
import csv
import zipfile
import openpyxl
import docx
from datetime import datetime
from urllib.parse import urlparse, urljoin
from bs4 import BeautifulSoup
from requests_html import HTMLSession
from transformers import BertTokenizer
from PyPDF2 import PdfReader
import html2text
from tqdm import tqdm
import json

# Configuración de registro para suprimir mensajes de advertencia innecesarios
import logging
logging.basicConfig(level=logging.ERROR)

# Lista de stopwords en español
STOPWORDS = set([
    "de", "la", "que", "el", "en", "y", "a", "los", "del", "se", "las", "por",
    "un", "para", "con", "no", "una", "su", "al", "es", "lo", "como", "más",
    "pero", "sus", "le", "ya", "o", "fue", "este", "ha", "sí", "porque",
    "esta", "son", "entre", "cuando", "muy", "sin", "sobre", "también", "me",
    "hasta", "hay", "donde", "quien", "desde", "todo", "nos", "durante",
    "todos", "uno", "les", "ni", "contra", "otros", "ese", "eso", "ante",
    "ellos", "e", "esto", "mí", "antes", "algunos", "qué", "unos", "yo",
    "otro", "otras", "otra", "él", "tanto", "esa", "estos", "mucho",
    "quienes", "nada", "muchos", "cual", "poco", "ella", "estar", "estas",
    "algunas", "algo", "nosotros", "mi", "mis", "tú", "te", "ti", "tu",
    "tus", "ellas", "nosotras", "vosotros", "vosotras", "os", "mío", "mía",
    "míos", "mías", "tuyo", "tuya", "tuyos", "tuyas", "suyo", "suya",
    "suyos", "suyas", "nuestro", "nuestra", "nuestros", "nuestras",
    "vuestro", "vuestra", "vuestros", "vuestras", "esos", "esas", "estoy",
    "estás", "está", "estamos", "estáis", "están", "esté", "estés",
    "estemos", "estéis", "estén", "estaré", "estarás", "estará",
    "estaremos", "estaréis", "estarán", "estaría", "estarías",
    "estaríamos", "estaríais", "estarían", "estaba", "estabas",
    "estábamos", "estabais", "estaban", "estuve", "estuviste", "estuvo",
    "estuvimos", "estuvisteis", "estuvieron", "estuviera", "estuvieras",
    "estuviéramos", "estuvierais", "estuvieran", "estuviese",
    "estuvieses", "estuviésemos", "estuvieseis", "estuviesen", "estando",
    "estado", "estada", "estados", "estadas", "estad"
])

# Inicializar el tokenizador de BERT en español
tokenizer = BertTokenizer.from_pretrained('dccuchile/bert-base-spanish-wwm-cased')

def clean_text(text):
    """
    Limpia el texto eliminando HTML, puntuación y stopwords.
    """
    # Eliminar etiquetas HTML
    soup = BeautifulSoup(text, 'html.parser')
    text = soup.get_text(separator=" ")

    # Convertir a minúsculas
    text = text.lower()

    # Eliminar URLs
    text = re.sub(r'http\S+', '', text)

    # Eliminar caracteres especiales y números
    text = re.sub(r'[^a-záéíóúñü\s]', '', text)

    # Eliminar espacios extra
    text = re.sub(r'\s+', ' ', text).strip()

    # Eliminar stopwords
    words = text.split()
    filtered_words = [word for word in words if word not in STOPWORDS]
    cleaned_text = ' '.join(filtered_words)

    return cleaned_text

def tokenize_and_save(text, filename, destination_folder):
    """
    Tokeniza el texto utilizando BERT y guarda los tokens en un archivo.
    """
    tokens = tokenizer.encode(
        text,
        truncation=True,
        max_length=512,
        add_special_tokens=True
    )
    tokens_str = ' '.join(map(str, tokens))
    short_filename = generate_short_filename(filename)
    short_file_path = os.path.join(destination_folder, short_filename)
    with open(short_file_path, 'w', encoding='utf-8') as f:
        f.write(tokens_str)

def generate_short_filename(filename):
    """
    Genera un nombre de archivo único utilizando un hash SHA256.
    """
    hash_name = hashlib.sha256(filename.encode()).hexdigest()
    short_name = hash_name[:10]
    return f"{short_name}_tokenized.txt"

def tokenize_all_articles(articles_folder, destination_folder):
    """
    Tokeniza todos los artículos en la carpeta especificada.
    """
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    print("Iniciando proceso de tokenización...")
    total_articles = 0
    total_size = 0

    for root, dirs, files in os.walk(articles_folder):
        for file in files:
            if file.endswith('.txt'):
                file_path = os.path.join(root, file)
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    tokenize_and_save(content, file, destination_folder)
                    total_articles += 1
                    total_size += os.path.getsize(file_path)

    total_size_mb = total_size / (1024 * 1024)
    print(f"Tokenización completada para {total_articles} artículos.")
    print(f"Tamaño total de artículos tokenizados: {total_size_mb:.2f} MB.")

def read_pdf(pdf_path):
    """
    Lee y extrae texto de un archivo PDF.
    """
    content = ''
    try:
        with open(pdf_path, 'rb') as f:
            pdf_reader = PdfReader(f)
            for page in pdf_reader.pages:
                text = page.extract_text()
                if text:
                    content += text + '\n'
    except Exception as e:
        print(f"Error al leer PDF {pdf_path}: {e}")
    return content

def read_csv(csv_path):
    """
    Lee y extrae texto de un archivo CSV.
    """
    content = ''
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                content += ' '.join(row) + '\n'
    except Exception as e:
        print(f"Error al leer CSV {csv_path}: {e}")
    return content

def read_docx(docx_path):
    """
    Lee y extrae texto de un archivo DOCX.
    """
    content = ''
    try:
        doc = docx.Document(docx_path)
        for paragraph in doc.paragraphs:
            content += paragraph.text + '\n'
    except Exception as e:
        print(f"Error al leer DOCX {docx_path}: {e}")
    return content

def read_xlsx(xlsx_path):
    """
    Lee y extrae texto de un archivo XLSX.
    """
    content = ''
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                row_text = ' '.join([str(cell.value) if cell.value is not None else '' for cell in row])
                content += row_text + '\n'
    except Exception as e:
        print(f"Error al leer XLSX {xlsx_path}: {e}")
    return content

def read_zip(zip_path):
    """
    Lee y extrae texto de un archivo ZIP.
    """
    content = ''
    try:
        with zipfile.ZipFile(zip_path, 'r') as z:
            for filename in z.namelist():
                with z.open(filename) as f:
                    file_content = f.read().decode('utf-8', errors='ignore')
                    content += file_content + '\n'
    except Exception as e:
        print(f"Error al leer ZIP {zip_path}: {e}")
    return content

def format_content(html_content):
    """
    Convierte contenido HTML a texto plano.
    """
    h = html2text.HTML2Text()
    h.ignore_links = True
    h.ignore_images = True
    text = h.handle(html_content)
    return text

def process_files(files_folder, destination_folder):
    """
    Procesa y tokeniza todos los archivos en la carpeta especificada.
    """
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    print("Procesando archivos descargados...")
    total_files = 0
    total_size = 0

    for root, dirs, files in os.walk(files_folder):
        for file in files:
            file_path = os.path.join(root, file)
            content = ''

            if file.endswith('.pdf'):
                content = read_pdf(file_path)
            elif file.endswith('.csv'):
                content = read_csv(file_path)
            elif file.endswith('.txt'):
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                except Exception as e:
                    print(f"Error al leer TXT {file_path}: {e}")
            elif file.endswith('.docx'):
                content = read_docx(file_path)
            elif file.endswith('.xlsx'):
                content = read_xlsx(file_path)
            elif file.endswith('.zip'):
                content = read_zip(file_path)
            elif file.endswith('.html') or file.endswith('.md'):
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        html_content = f.read()
                        content = format_content(html_content)
                except Exception as e:
                    print(f"Error al leer HTML/MD {file_path}: {e}")
            else:
                print(f"Formato de archivo no soportado: {file}")
                continue

            if content:
                cleaned_text = clean_text(content)
                tokenize_and_save(cleaned_text, file, destination_folder)
                total_files += 1
                total_size += os.path.getsize(file_path)

    total_size_mb = total_size / (1024 * 1024)
    print(f"Procesamiento completado para {total_files} archivos.")
    print(f"Tamaño total de archivos procesados: {total_size_mb:.2f} MB.")

def download_and_save_file(url, destination_folder):
    """
    Descarga y guarda un archivo desde la URL especificada.
    """
    try:
        print(f"Descargando archivo: {url}")
        response = requests.get(url, stream=True, timeout=30)
        if response.status_code == 200:
            filename = clean_filename(url.split('/')[-1])
            if not filename:
                filename = 'archivo_descargado'
            file_path = os.path.join(destination_folder, filename)
            with open(file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            print(f"Archivo descargado: {file_path}")
        else:
            print(f"Error al descargar {url}: Código de estado {response.status_code}")
    except Exception as e:
        print(f"Error al descargar {url}: {e}")

def extract_and_save_article(url, articles_folder):
    """
    Extrae y guarda el contenido de un artículo desde la URL especificada.
    """
    try:
        print(f"Extrayendo artículo: {url}")
        response = requests.get(url, timeout=30)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')

            # Obtener título
            title_tag = soup.find('title')
            title = title_tag.get_text().strip() if title_tag else None

            # Obtener contenido
            paragraphs = soup.find_all('p')
            content = ' '.join([para.get_text() for para in paragraphs])

            if content.strip():
                cleaned_text = clean_text(content)
                if title:
                    filename = clean_filename(title) + '.txt'
                else:
                    parsed_url = urlparse(url)
                    filename = clean_filename(parsed_url.path.split('/')[-1]) + '.txt'

                file_path = os.path.join(articles_folder, filename)

                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(cleaned_text)

                print(f"Artículo guardado: {file_path}")
            else:
                print(f"No se encontró contenido en {url}")
        else:
            print(f"Error al acceder a {url}: Código de estado {response.status_code}")
    except Exception as e:
        print(f"Error al extraer artículo de {url}: {e}")

def get_page_title(url):
    """
    Obtiene el título de la página web desde la URL especificada.
    """
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            title_tag = soup.find('title')
            return title_tag.get_text().strip() if title_tag else None
        else:
            return None
    except Exception as e:
        print(f"Error al obtener el título de la página {url}: {e}")
        return None

def clean_filename(name):
    """
    Limpia el nombre del archivo eliminando caracteres no permitidos.
    """
    if name is None:
        return 'sin_nombre'
    name = re.sub(r'[\\/*?:"<>|]', "_", name)
    name = re.sub(r'\s+', '_', name)
    return name[:100]  # Limitar la longitud del nombre

def register_processed_notifications(base_folder, urls):
    """
    Registra las URLs ya procesadas para evitar duplicados.
    """
    if not os.path.exists(base_folder):
        os.makedirs(base_folder)

    txt_path = os.path.join(base_folder, "processed_articles.txt")
    processed_urls = set()

    if os.path.exists(txt_path):
        with open(txt_path, 'r') as f:
            processed_urls = set(f.read().splitlines())

    urls_to_process = [url for url in urls if url not in processed_urls]

    with open(txt_path, 'a') as f:
        for url in urls_to_process:
            f.write(url + "\n")

    if processed_urls:
        print(f"Artículos ya procesados: {len(processed_urls)}")
    else:
        print("No hay artículos procesados previamente.")

    return urls_to_process

def explore_wayback_machine(url, articles_folder):
    """
    Explora la Wayback Machine para obtener versiones archivadas de la URL.
    """
    try:
        print(f"Explorando Wayback Machine para: {url}")
        api_url = f"http://archive.org/wayback/available?url={url}"
        response = requests.get(api_url, timeout=10)
        data = response.json()

        if 'archived_snapshots' in data and 'closest' in data['archived_snapshots']:
            archive_url = data['archived_snapshots']['closest']['url']
            print(f"Descargando desde Wayback Machine: {archive_url}")
            extract_and_save_article(archive_url, articles_folder)
        else:
            print(f"No se encontró versión archivada para {url}")
    except Exception as e:
        print(f"Error al explorar Wayback Machine para {url}: {e}")

def get_folder_info(path):
    """
    Obtiene información de la carpeta: tamaño total y número de archivos.
    """
    total_size = 0
    total_files = 0
    for dirpath, dirnames, filenames in os.walk(path):
        for f in filenames:
            fp = os.path.join(dirpath, f)
            total_size += os.path.getsize(fp)
            total_files += 1
    return total_size, total_files

def explore_and_extract_articles(url, articles_folder, files_folder, processed_urls, size_limit, depth=0, max_depth=2):
    """
    Explora y extrae artículos y archivos desde la URL especificada.
    """
    if depth > max_depth:
        return

    print(f"Explorando {url} en profundidad {depth}...")
    try:
        session = HTMLSession()
        response = session.get(url, timeout=30)
        response.html.render(timeout=30, sleep=1)
        links = response.html.absolute_links
        session.close()
    except Exception as e:
        print(f"Error al acceder a {url}: {e}")
        return

    for link in links:
        if link in processed_urls:
            continue

        processed_urls.add(link)

        parsed_link = urlparse(link)
        file_extension = os.path.splitext(parsed_link.path)[1].lower()

        if file_extension in ['.pdf', '.csv', '.txt', '.xlsx', '.docx', '.html', '.md', '.zip']:
            download_and_save_file(link, files_folder)
        elif 'mailto:' in link or 'tel:' in link:
            continue
        else:
            extract_and_save_article(link, articles_folder)
            # Exploración recursiva limitada por profundidad
            explore_and_extract_articles(link, articles_folder, files_folder, processed_urls, size_limit, depth + 1, max_depth)

        # Comprobar el tamaño total de las carpetas
        total_size_articles, _ = get_folder_info(articles_folder)
        total_size_files, _ = get_folder_info(files_folder)
        total_size = total_size_articles + total_size_files

        if total_size >= size_limit:
            print("Se ha alcanzado el límite de tamaño de 50 GB. Deteniendo exploración.")
            return

def main():
    # URLs de los sitios web de noticias a explorar
    urls = [
        'https://reactionary.international/database/',
        'https://aleph.occrp.org/',
        'https://offshoreleaks.icij.org/',
        'https://www.publico.es/',
        'https://www.elsaltodiario.com/',
        'https://www.nytimes.com/',
        'https://www.theguardian.com/',
        'https://www.lemonde.fr/',
        'https://www.spiegel.de/',
        'https://elpais.com/',
        'https://www.repubblica.it/',
        'https://www.scmp.com/',
        'https://www.smh.com.au/',
        'https://www.globo.com/',
        'https://timesofindia.indiatimes.com/',
        'https://www.asahi.com/',
        'https://www.washingtonpost.com/',
        'https://www.aljazeera.com/',
        'https://www.folha.uol.com.br/',
        'https://www.telegraph.co.uk/',
        'https://www.corriere.it/',
        'https://www.clarin.com/',
        'https://www.eluniversal.com.mx/',
        'https://www.welt.de/',
        'https://www.lanacion.com.ar/'
    ]

    # Configuración de carpetas
    base_folder = '/var/www/flujos/FLUJOS_DATOS/NOTICIAS'
    articles_folder = os.path.join(base_folder, 'articulos')
    files_folder = os.path.join(base_folder, 'archivos')
    tokenized_folder = os.path.join(base_folder, 'tokenized')

    # Crear carpetas si no existen
    for folder in [articles_folder, files_folder, tokenized_folder]:
        if not os.path.exists(folder):
            os.makedirs(folder)

    # Límite de tamaño de carpeta en bytes (50 GB)
    FOLDER_SIZE_LIMIT = 50 * 1024 * 1024 * 1024  # 50 GB

    # Registrar URLs ya procesadas
    urls_to_process = register_processed_notifications(base_folder, urls)

    # Conjunto para almacenar las URLs ya procesadas
    processed_urls = set()

    # Explorar y extraer artículos y archivos
    for url in urls_to_process:
        print(f"\nProcesando URL: {url}")
        explore_and_extract_articles(url, articles_folder, files_folder, processed_urls, FOLDER_SIZE_LIMIT)
        explore_wayback_machine(url, articles_folder)

    # Procesar y tokenizar archivos descargados
    process_files(files_folder, tokenized_folder)

    # Tokenizar todos los artículos
    tokenize_all_articles(articles_folder, tokenized_folder)

    # Obtener información de las carpetas
    total_size_articles, total_files_articles = get_folder_info(articles_folder)
    total_size_files, total_files_files = get_folder_info(files_folder)
    total_size_tokenized, total_files_tokenized = get_folder_info(tokenized_folder)

    print("\nResumen del proceso:")
    print(f"Artículos descargados: {total_files_articles}")
    print(f"Tamaño total de artículos: {total_size_articles / (1024 * 1024):.2f} MB")
    print(f"Archivos descargados: {total_files_files}")
    print(f"Tamaño total de archivos: {total_size_files / (1024 * 1024):.2f} MB")
    print(f"Archivos tokenizados: {total_files_tokenized}")
    print(f"Tamaño total de archivos tokenizados: {total_size_tokenized / (1024 * 1024):.2f} MB")

if __name__ == "__main__":
    main()
